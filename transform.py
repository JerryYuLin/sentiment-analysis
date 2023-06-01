from opencc import OpenCC
import csv
import openpyxl

workbook = openpyxl.load_workbook("data.xlsx")
sheet = workbook.worksheets[0]

with open("new_data.csv", "w", encoding="utf-8", newline="") as write:
    cc = OpenCC("s2twp")

    writer = csv.writer(write)

    for i in range(1, sheet.max_row + 1):
        try:
            writer.writerow(
                [
                    cc.convert(sheet.cell(row=i, column=1).value),
                    sheet.cell(row=i, column=7).value,
                ]
            )
        except:
            pass
